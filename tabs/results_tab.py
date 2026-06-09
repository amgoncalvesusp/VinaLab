# -*- coding: utf-8 -*-
"""Results table, chart, exports, and live log tab."""

import json
import logging
import os
from pathlib import Path
import shutil
import subprocess
import statistics
import sys
import tempfile

import pandas as pd
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
from matplotlib.figure import Figure
from openpyxl.styles import Font, PatternFill
from PySide6.QtCore import QPoint, QUrl, Signal, Qt
from PySide6.QtGui import QAction, QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDoubleSpinBox,
    QFileDialog,
    QHeaderView,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMenu,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QSlider,
    QSpinBox,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QToolButton,
    QVBoxLayout,
    QWidget,
)

try:
    from PySide6.QtWebEngineWidgets import QWebEngineView
except ImportError:  # pragma: no cover - installer should provide PySide6-WebEngine
    QWebEngineView = None

from core.docking_engine import (
    convert_with_obabel,
    extract_pose_model,
    find_obabel_executable,
)
from core.file_utils import clean_pdbqt_text
from core.i18n import I18n
from core.rmsd import (
    pdbqt_heavy_coordinates,
    pose_pair_rmsd,
    symmetry_corrected_rmsd,
)
from core.scrolling import ScrollManager
from tabs.results_clustering import build_pose_clusters
from tabs.results_plotly import (
    affinity_chart_html,
    cluster_chart_html,
    plotly_available,
)
from tabs.results_dialogs import (
    ComparisonDialog,
    ExportComplexDialog,
    ProtocolValidationDialog,
)
from tabs.results_view import (
    apply_header_tooltips,
    build_pose_view_html,
    prepare_pose_view_files,
    safe_export_name,
)

logger = logging.getLogger(__name__)

NO_WINDOW = subprocess.CREATE_NO_WINDOW if sys.platform.startswith("win") else 0


class ResultsTab(QWidget):
    """Display docking results and live docking logs."""

    results_changed = Signal(list)
    chart_updated = Signal(object)

    def __init__(self) -> None:
        """Create the results table, chart, export buttons, and log console."""
        super().__init__()
        self.lang = "pt"
        self.results: list[dict] = []
        self.filtered_results: list[dict] = []
        self.table_state: dict = {"pins": {}, "notes": {}, "filters": {}}
        self.state_path: Path | None = None
        self._state_loaded_for: Path | None = None
        self._rendering_table = False
        self._sort_criteria: list[tuple[str, bool]] = []
        self.chart_path: Path | None = None
        self.table = QTableWidget(0, 0)
        self.log_console = QTextEdit()
        self.figure = Figure(figsize=(8, 3), tight_layout=True)
        self.canvas = FigureCanvasQTAgg(self.figure)
        # Interactive Plotly chart (hover shows molecule); matplotlib stays as the
        # PNG source for the PDF report and as a fallback when WebEngine/Plotly are
        # unavailable.
        self._charts_interactive = QWebEngineView is not None and plotly_available()
        self.affinity_web = QWebEngineView() if self._charts_interactive else None
        self.excel_button = QPushButton()
        self.csv_button = QPushButton()
        self.filtered_export_button = QPushButton()
        self.view_button = QPushButton()
        self.compare_button = QPushButton()
        self.export_complex_button = QPushButton()
        self.ligand_filter = QLineEdit()
        self.scoring_filter_button = QToolButton()
        self.scoring_filter_menu = QMenu(self)
        self.scoring_filter_actions: dict[str, QAction] = {}
        self.affinity_min = self._range_spin(-9999.0)
        self.affinity_max = self._range_spin(9999.0)
        self.rmsd_min = self._range_spin(0.0)
        self.rmsd_max = self._range_spin(9999.0)
        self.pinned_only = QCheckBox()
        self.preview_status = QLabel()
        self.pose_detail_label = QLabel()
        self.preview_view = QWebEngineView() if QWebEngineView is not None else None
        self.preview_tabs = QTabWidget()
        self.interaction_cutoff = QComboBox()
        self.interaction_table = QTableWidget(0, 7)
        self.export_interactions_button = QPushButton()
        self.current_interactions: list[dict] = []
        self.current_interaction_row: dict | None = None
        self.current_box: dict | None = None
        self.current_preview_row: dict | None = None
        self.current_receptor_pdb: Path | None = None
        self.current_pose_pdb: Path | None = None
        self.current_receptor_path: Path | None = None
        self.consensus_widget = QWidget()
        self.consensus_table = QTableWidget(0, 0)
        self.consensus_tab_index = -1
        self.clusters_widget = QWidget()
        self.cluster_table = QTableWidget(0, 6)
        self.cluster_cutoff_slider = QSlider(Qt.Horizontal)
        self.cluster_cutoff_label = QLabel()
        self.cluster_export_button = QPushButton()
        self.cluster_figure = Figure(figsize=(4, 2), tight_layout=True)
        self.cluster_canvas = FigureCanvasQTAgg(self.cluster_figure)
        self.cluster_web = QWebEngineView() if self._charts_interactive else None
        self.cluster_tab_index = -1
        self.current_clusters: list[dict] = []
        self.column_keys: list[str] = []
        self.consensus_plot_widget = QWidget()
        self.consensus_plot_x_combo = QComboBox()
        self.consensus_plot_y_combo = QComboBox()
        self.consensus_plot_z_combo = QComboBox()
        self.consensus_plot_3d_checkbox = QCheckBox()
        self.consensus_plot_refresh_button = QPushButton()
        self.consensus_plot_figure = Figure(figsize=(5, 4), tight_layout=True)
        self.consensus_plot_canvas = FigureCanvasQTAgg(self.consensus_plot_figure)
        self.consensus_plot_tab_index = -1
        self._build_ui()
        self.retranslate_ui(self.lang)

    def clear_results(self) -> None:
        """Clear existing table rows, chart, logs, and stored results."""
        self.results = []
        self.filtered_results = []
        self.table.setRowCount(0)
        self.consensus_table.setRowCount(0)
        self._set_consensus_tab_visible(False)
        self.cluster_table.setRowCount(0)
        self._set_clusters_tab_visible(False)
        self.current_clusters = []
        self.log_console.clear()
        self.figure.clear()
        self.canvas.draw_idle()
        self.results_changed.emit([])

    def add_results(self, rows: list[dict]) -> None:
        """Append result rows to the table and refresh the chart."""
        self.results.extend(rows)
        self._sync_state_path()
        self._refresh_scoring_filter_options()
        self._apply_filters()
        self._update_consensus_tab()
        self._update_clusters_tab()
        self._update_chart()
        self._update_consensus_plot()
        self.results_changed.emit(list(self.results))

    def append_log(self, message: str) -> None:
        """Append a log message from the docking worker."""
        self.log_console.append(message.rstrip())

    def open_export_dialog(self) -> None:
        """Open the complex-export dialog."""
        if not self.results:
            QMessageBox.information(
                self,
                I18n.get("export_complex", self.lang),
                I18n.get("rt_no_poses_export", self.lang),
            )
            return
        dialog = ExportComplexDialog(self.results, self, self.lang)
        dialog.exec()

    def export_to_csv(self) -> None:
        """Export the full results table to CSV."""
        if not self.results:
            return
        file_name, _ = QFileDialog.getSaveFileName(
            self, I18n.get("export_csv_dialog", self.lang), "", "Arquivos CSV (*.csv)"
        )
        if file_name:
            self._dataframe().to_csv(Path(file_name), index=False)

    def export_to_excel(self) -> None:
        """Export the full results table to Excel with basic formatting."""
        if not self.results:
            return
        file_name, _ = QFileDialog.getSaveFileName(
            self, I18n.get("export_excel_dialog", self.lang), "", "Excel Files (*.xlsx)"
        )
        if not file_name:
            return
        output_path = Path(file_name)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            self._dataframe().to_excel(writer, sheet_name="Results", index=False)
            worksheet = writer.book["Results"]
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4C3A75")
            for row in range(2, worksheet.max_row + 1):
                affinity = worksheet.cell(row=row, column=4).value
                fill = self._excel_fill_for_affinity(float(affinity))
                for column in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=row, column=column).fill = fill
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = (
                    max_length + 2
                )

    def retranslate_ui(self, lang: str) -> None:
        """Retranslate results tab controls."""
        self.lang = lang
        self.excel_button.setText(I18n.get("export_excel", lang))
        self.csv_button.setText(I18n.get("export_csv", lang))
        self.filtered_export_button.setText(I18n.get("export_filtered_subset", lang))
        self.view_button.setText(I18n.get("view_selected_pose", lang))
        self.compare_button.setText(I18n.get("compare_button", lang))
        self.export_complex_button.setText(I18n.get("export_complex", lang))
        self.pinned_only.setText(I18n.get("pinned_only", lang))
        self.ligand_filter.setPlaceholderText(I18n.get("ligand_name_filter", lang))
        self.scoring_filter_button.setText(I18n.get("scoring_all", lang))
        self.preview_status.setText(I18n.get("select_pose_preview", lang))
        self.export_interactions_button.setText(
            I18n.get("export_interaction_table", lang)
        )
        self.cluster_export_button.setText(I18n.get("export_best_cluster", lang))
        self.preview_tabs.setTabText(0, I18n.get("results_table", lang))
        if self.preview_tabs.count() > 1:
            self.preview_tabs.setTabText(1, I18n.get("interactions", lang))
        if self.preview_tabs.count() > 2 and self.preview_view is not None:
            self.preview_tabs.setTabText(2, I18n.get("pose_3d_view", lang))
        consensus_index = self._side_tab_index(self.consensus_widget)
        if consensus_index >= 0:
            self.preview_tabs.setTabText(consensus_index, I18n.get("consensus", lang))
        clusters_index = self._side_tab_index(self.clusters_widget)
        if clusters_index >= 0:
            self.preview_tabs.setTabText(clusters_index, I18n.get("clusters", lang))
        self.excel_button.setToolTip(I18n.get("export_excel", lang))
        self.csv_button.setToolTip(I18n.get("export_csv", lang))
        self.filtered_export_button.setToolTip(I18n.get("rt_tip_export_filtered", lang))
        self.view_button.setToolTip(I18n.get("rt_tip_view_pymol", lang))
        self.compare_button.setToolTip(I18n.get("rt_tip_compare", lang))
        self.export_complex_button.setToolTip(I18n.get("rt_tip_export_complex", lang))
        self.export_interactions_button.setToolTip(
            I18n.get("rt_tip_export_interactions", lang)
        )
        self.cluster_export_button.setToolTip(I18n.get("rt_tip_cluster_export", lang))
        self._refresh_scoring_filter_button()
        self._update_chart()

    def _build_ui(self) -> None:
        """Build the results tab layout."""
        layout = QVBoxLayout(self)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSortingEnabled(False)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.cellClicked.connect(
            lambda _row, _column: self.load_selected_pose_preview()
        )
        self.table.cellDoubleClicked.connect(
            lambda _row, _column: self.view_selected_pose()
        )
        self.table.itemChanged.connect(self._table_item_changed)
        self.table.horizontalHeader().sectionClicked.connect(self._header_clicked)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._open_context_menu)
        ScrollManager.optimize(self.table)
        self.log_console.setReadOnly(True)
        self.log_console.setObjectName("log_console")
        ScrollManager.optimize(self.log_console)

        filter_row = QHBoxLayout()
        self.ligand_filter.textChanged.connect(self._apply_filters)
        self.scoring_filter_button.setMenu(self.scoring_filter_menu)
        self.scoring_filter_button.setPopupMode(QToolButton.InstantPopup)
        self.affinity_min.valueChanged.connect(self._apply_filters)
        self.affinity_max.valueChanged.connect(self._apply_filters)
        self.rmsd_min.valueChanged.connect(self._apply_filters)
        self.rmsd_max.valueChanged.connect(self._apply_filters)
        self.pinned_only.toggled.connect(self._apply_filters)
        filter_row.addWidget(QLabel(I18n.get("ligand_col", self.lang)))
        filter_row.addWidget(self.ligand_filter, stretch=2)
        filter_row.addWidget(self.scoring_filter_button)
        filter_row.addWidget(QLabel(I18n.get("affinity_col", self.lang)))
        filter_row.addWidget(self.affinity_min)
        filter_row.addWidget(self.affinity_max)
        filter_row.addWidget(QLabel("RMSD"))
        filter_row.addWidget(self.rmsd_min)
        filter_row.addWidget(self.rmsd_max)
        filter_row.addWidget(self.pinned_only)

        results_panel = QWidget()
        results_layout = QVBoxLayout(results_panel)
        results_layout.setContentsMargins(0, 0, 0, 0)
        results_layout.addLayout(filter_row)
        results_layout.addWidget(self.table)

        button_row = QHBoxLayout()
        self.excel_button.clicked.connect(self.export_to_excel)
        self.csv_button.clicked.connect(self.export_to_csv)
        self.filtered_export_button.clicked.connect(self.export_filtered_subset)
        self.view_button.clicked.connect(self.view_selected_pose)
        self.compare_button.clicked.connect(self.open_comparison_dialog)
        button_row.addWidget(self.view_button)
        button_row.addWidget(self.compare_button)
        button_row.addWidget(self.filtered_export_button)
        button_row.addWidget(self.excel_button)
        button_row.addWidget(self.csv_button)
        button_row.addStretch()
        affinity_widget = (
            self.affinity_web if self.affinity_web is not None else self.canvas
        )
        affinity_widget.setMinimumHeight(360)
        results_layout.addWidget(affinity_widget, stretch=3)
        results_layout.addLayout(button_row)
        results_layout.addWidget(self.log_console, stretch=1)

        interaction_widget = QWidget()
        interaction_layout = QVBoxLayout(interaction_widget)
        interaction_layout.setContentsMargins(0, 0, 0, 0)
        interaction_controls = QHBoxLayout()
        self.interaction_cutoff.addItems(["4", "5", "6"])
        self.interaction_cutoff.setCurrentText("4")
        self.interaction_cutoff.currentTextChanged.connect(
            lambda _value: self._refresh_interactions_for_selected_pose()
        )
        self.export_interactions_button.clicked.connect(self.export_interaction_table)
        interaction_controls.addWidget(QLabel(I18n.get("contact_cutoff", self.lang)))
        interaction_controls.addWidget(self.interaction_cutoff)
        interaction_controls.addWidget(self.export_interactions_button)
        interaction_controls.addStretch()
        self.interaction_table.setHorizontalHeaderLabels(
            [
                I18n.get("residue", self.lang),
                I18n.get("interaction_type", self.lang),
                I18n.get("donor", self.lang),
                I18n.get("acceptor", self.lang),
                I18n.get("distance_a", self.lang),
                I18n.get("angle", self.lang),
                I18n.get("frequency_top10", self.lang),
            ]
        )
        apply_header_tooltips(self.interaction_table)
        self.interaction_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )
        self.interaction_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        ScrollManager.optimize(self.interaction_table)
        interaction_layout.addLayout(interaction_controls)
        interaction_layout.addWidget(self.interaction_table, stretch=1)

        consensus_layout = QVBoxLayout(self.consensus_widget)
        consensus_layout.setContentsMargins(0, 0, 0, 0)
        consensus_hint = QLabel(I18n.get("rt_consensus_hint", self.lang))
        consensus_hint.setObjectName("label_muted")
        self.consensus_table.setSortingEnabled(True)
        self.consensus_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )
        self.consensus_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        ScrollManager.optimize(self.consensus_table)
        consensus_layout.addWidget(consensus_hint)
        consensus_layout.addWidget(self.consensus_table, stretch=1)

        consensus_plot_layout = QVBoxLayout(self.consensus_plot_widget)
        consensus_plot_layout.setContentsMargins(0, 0, 0, 0)
        consensus_plot_controls = QHBoxLayout()
        consensus_plot_controls.addWidget(QLabel("X:"))
        consensus_plot_controls.addWidget(self.consensus_plot_x_combo)
        consensus_plot_controls.addWidget(QLabel("Y:"))
        consensus_plot_controls.addWidget(self.consensus_plot_y_combo)
        self.consensus_plot_3d_checkbox.setText("3D")
        consensus_plot_controls.addWidget(self.consensus_plot_3d_checkbox)
        consensus_plot_controls.addWidget(QLabel("Z:"))
        consensus_plot_controls.addWidget(self.consensus_plot_z_combo)
        self.consensus_plot_z_combo.setEnabled(False)
        self.consensus_plot_refresh_button.setText("Atualizar plot")
        consensus_plot_controls.addWidget(self.consensus_plot_refresh_button)
        consensus_plot_controls.addStretch()
        self.consensus_plot_3d_checkbox.toggled.connect(
            self.consensus_plot_z_combo.setEnabled
        )
        self.consensus_plot_3d_checkbox.toggled.connect(
            lambda _checked: self._update_consensus_plot()
        )
        self.consensus_plot_refresh_button.clicked.connect(self._update_consensus_plot)
        self.consensus_plot_x_combo.currentIndexChanged.connect(
            lambda _index: self._update_consensus_plot()
        )
        self.consensus_plot_y_combo.currentIndexChanged.connect(
            lambda _index: self._update_consensus_plot()
        )
        self.consensus_plot_z_combo.currentIndexChanged.connect(
            lambda _index: self._update_consensus_plot()
        )
        consensus_plot_layout.addLayout(consensus_plot_controls)
        consensus_plot_layout.addWidget(self.consensus_plot_canvas, stretch=1)

        clusters_layout = QVBoxLayout(self.clusters_widget)
        clusters_layout.setContentsMargins(0, 0, 0, 0)
        cluster_controls = QHBoxLayout()
        self.cluster_cutoff_slider.setRange(50, 500)
        self.cluster_cutoff_slider.setValue(200)
        self.cluster_cutoff_slider.valueChanged.connect(self._cluster_cutoff_changed)
        self.cluster_export_button.clicked.connect(self.export_cluster_representatives)
        self.cluster_cutoff_label.setText("RMSD cutoff: 2.0 Å")
        cluster_controls.addWidget(self.cluster_cutoff_label)
        cluster_controls.addWidget(self.cluster_cutoff_slider, stretch=1)
        cluster_controls.addWidget(self.cluster_export_button)
        self.cluster_table.setHorizontalHeaderLabels(
            [
                I18n.get("ligand_col", self.lang),
                I18n.get("cluster_id", self.lang),
                I18n.get("cluster_size", self.lang),
                I18n.get("best_score", self.lang),
                I18n.get("representative_pose", self.lang),
                I18n.get("members", self.lang),
            ]
        )
        apply_header_tooltips(self.cluster_table)
        self.cluster_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.cluster_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.cluster_table.cellClicked.connect(self._cluster_row_clicked)
        ScrollManager.optimize(self.cluster_table)
        clusters_layout.addLayout(cluster_controls)
        clusters_layout.addWidget(self.cluster_table, stretch=2)
        cluster_chart_widget = (
            self.cluster_web if self.cluster_web is not None else self.cluster_canvas
        )
        cluster_chart_widget.setMinimumHeight(320)
        clusters_layout.addWidget(cluster_chart_widget, stretch=3)

        self.preview_tabs.addTab(results_panel, I18n.get("results_table", self.lang))
        self.preview_tabs.addTab(
            interaction_widget, I18n.get("interactions", self.lang)
        )
        if self.preview_view is not None:
            self.preview_tabs.addTab(
                self.preview_view, I18n.get("pose_3d_view", self.lang)
            )

        analysis_panel = QWidget()
        analysis_layout = QVBoxLayout(analysis_panel)
        analysis_layout.setContentsMargins(0, 0, 0, 0)
        self.pose_detail_label.setObjectName("label_muted")
        self.pose_detail_label.setWordWrap(True)
        self.pose_detail_label.setText("Nenhuma pose selecionada.")
        self.preview_status.setObjectName("label_muted")
        self.preview_status.setWordWrap(True)
        self.export_complex_button.clicked.connect(self.open_export_dialog)
        analysis_layout.addWidget(self.preview_tabs, stretch=1)
        analysis_layout.addWidget(self.preview_status)
        analysis_layout.addWidget(self.pose_detail_label)
        analysis_layout.addWidget(self.export_complex_button)

        layout.addWidget(analysis_panel, stretch=1)

    def _populate_table(self) -> None:
        """Render stored result rows into the QTableWidget."""
        scoring_headers = self._scoring_headers()
        headers = [
            I18n.get("ligand_col", self.lang),
            I18n.get("pose_rank", self.lang),
            I18n.get("docking_score", self.lang),
            *scoring_headers,
            I18n.get("rmsd_best_pose", self.lang),
            I18n.get("pinned", self.lang),
            I18n.get("notes", self.lang),
            I18n.get("scoring_error", self.lang),
        ]
        self.column_keys = [
            "ligand_name",
            "pose_rank",
            "docking_score",
            *[f"score::{header}" for header in scoring_headers],
            "rmsd_lb",
            "pinned",
            "notes",
            "scoring_error",
        ]
        self._rendering_table = True
        self.table.clear()
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        apply_header_tooltips(self.table)
        self.table.setRowCount(len(self.filtered_results))
        for row_index, row in enumerate(self.filtered_results):
            values = self._table_values(row, scoring_headers)
            background = self._color_for_affinity(float(row.get("affinity", 0.0)))
            row_id = self._row_id(row)
            for column_index, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                item.setData(Qt.UserRole, row_id)
                key = self.column_keys[column_index]
                if key == "pinned":
                    item.setFlags(
                        item.flags()
                        | Qt.ItemIsUserCheckable
                        | Qt.ItemIsEnabled
                        | Qt.ItemIsSelectable
                    )
                    item.setCheckState(
                        Qt.Checked if self._row_pinned(row) else Qt.Unchecked
                    )
                    item.setText("")
                elif key == "notes":
                    item.setFlags(
                        item.flags()
                        | Qt.ItemIsEditable
                        | Qt.ItemIsEnabled
                        | Qt.ItemIsSelectable
                    )
                else:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if self._numeric_column(key) and value != "":
                    item.setData(
                        Qt.DisplayRole,
                        float(value) if key != "pose_rank" else int(value),
                    )
                item.setBackground(background)
                self.table.setItem(row_index, column_index, item)
        self._rendering_table = False

    def _update_chart(self) -> None:
        """Refresh the embedded affinity distribution bar chart."""
        self.figure.clear()
        axis = self.figure.add_subplot(111)
        rows = self.filtered_results if self.filtered_results else self.results
        if rows:
            labels = [
                f"{row['ligand_name']} {row.get('scoring_function', row.get('scoring_key', ''))} #{row['mode']}"
                for row in rows
            ]
            affinities = [float(row["affinity"]) for row in rows]
            axis.bar(range(len(affinities)), affinities, color="#7c5cbf")
            axis.set_xticks(range(len(labels)))
            axis.set_xticklabels(labels, rotation=45, ha="right", fontsize=8)
            axis.set_ylabel(I18n.get("affinity_col", self.lang))
            axis.set_title(I18n.get("chart_title", self.lang))
        self.canvas.draw_idle()
        self.chart_path = Path(tempfile.gettempdir()) / "vinalab_affinity_chart.png"
        self.figure.savefig(self.chart_path, dpi=160, bbox_inches="tight")
        self.chart_updated.emit(self.chart_path)
        if self.affinity_web is not None and rows:
            try:
                html_path = affinity_chart_html(
                    rows,
                    I18n.get("affinity_col", self.lang),
                    I18n.get("chart_title", self.lang),
                )
                self.affinity_web.setUrl(QUrl.fromLocalFile(str(html_path)))
            except Exception as exc:  # noqa: BLE001 - chart is non-critical
                logger.warning(
                    "Falha ao renderizar gráfico Plotly de afinidade: %s", exc
                )

    def _dataframe(self, rows: list[dict] | None = None) -> pd.DataFrame:
        """Return results as a pandas DataFrame."""
        rows = self.results if rows is None else rows
        return pd.DataFrame(
            [
                (
                    {
                        "Ligand name": row["ligand_name"],
                        "Scoring": row.get(
                            "scoring_function", row.get("scoring_key", "")
                        ),
                        "Mode #": row["mode"],
                        "Affinity (kcal/mol)": row["affinity"],
                        "Vina affinity (kcal/mol)": row.get("vina_affinity", ""),
                        "RMSD l.b.": row["rmsd_lb"],
                        "RMSD u.b.": row["rmsd_ub"],
                        "Scoring error": row.get("scoring_error", ""),
                        "Pinned": self._row_pinned(row),
                        "Notes": self._row_note(row),
                        "CNNscore": row.get("cnn_score", ""),
                        "CNNaffinity": row.get("cnn_affinity", ""),
                        "Output file": row.get("output_file", ""),
                    }
                    | {
                        header: row.get("affinity", "")
                        if row.get("scoring_function", row.get("scoring_key", ""))
                        == header
                        else ""
                        for header in self._scoring_headers()
                    }
                )
                for row in rows
            ]
        )

    def _update_gnina_columns(self) -> None:
        """Show GNINA CNN columns only when GNINA values are present."""
        has_gnina = any(
            "cnn_score" in row or "cnn_affinity" in row for row in self.results
        )
        self.table.setColumnHidden(8, not has_gnina)
        self.table.setColumnHidden(9, not has_gnina)

    def _update_headers(self) -> None:
        """Set headers, including dynamic post-scoring columns."""
        dynamic_headers = self._scoring_headers()
        self.table.setHorizontalHeaderLabels(
            [
                I18n.get("ligand_col", self.lang),
                "Scoring",
                I18n.get("mode_col", self.lang),
                I18n.get("affinity_col", self.lang),
                "Vina affinity",
                I18n.get("rmsd_lb_col", self.lang),
                I18n.get("rmsd_ub_col", self.lang),
                "Scoring error",
                I18n.get("col_cnn_score", self.lang),
                I18n.get("col_cnn_affinity", self.lang),
                *dynamic_headers,
            ]
        )

    def view_selected_pose(self) -> None:
        """Open the selected receptor-pose complex in PyMOL."""
        row = self._selected_result_row()
        if row is None:
            QMessageBox.information(
                self,
                I18n.get("rt_view_pose_title", self.lang),
                I18n.get("rt_select_pose", self.lang),
            )
            return
        pymol_exe = self._find_pymol()
        if pymol_exe is None:
            QMessageBox.warning(
                self,
                "PyMOL",
                I18n.get("rt_pymol_not_found", self.lang),
            )
            return
        try:
            output_file = Path(row.get("output_file", ""))
            receptor_file = Path(row.get("receptor_file", ""))
            if not output_file.is_absolute():
                output_file = output_file.resolve()
            if not receptor_file.is_absolute():
                receptor_file = receptor_file.resolve()
            if not output_file.exists():
                raise FileNotFoundError(
                    f"Arquivo de saída não encontrado: {output_file}"
                )
            view_dir = Path(tempfile.mkdtemp(prefix="vinalab_pymol_"))
            pose_name = (
                f"{safe_export_name(row['ligand_name'])}_pose{int(row['mode'])}.pdbqt"
            )
            pose_file = view_dir / pose_name
            pose_text = extract_pose_model(
                output_file, int(row["mode"]), include_model=False
            )
            pose_file.write_text(pose_text, encoding="utf-8")
            cmd_args = [pymol_exe]
            if receptor_file.exists():
                cmd_args.append(str(receptor_file))
            cmd_args.append(str(pose_file))
            # No NO_WINDOW here: PyMOL is a GUI app and must show its window.
            subprocess.Popen(cmd_args)
        except Exception as exc:  # noqa: BLE001 - surface file/pose extraction errors to the user
            QMessageBox.critical(
                self,
                "PyMOL",
                I18n.get("rt_pymol_open_error", self.lang).format(exc=exc),
            )

    @staticmethod
    def _find_pymol() -> str | None:
        """Locate a PyMOL executable across PATH and common Windows installs.

        Schrodinger/Incentive PyMOL installs as PyMOLWin.exe / PyMOL.exe and is
        usually not on PATH, so shutil.which('pymol') alone failed silently.
        """
        for name in ("pymol", "pymol.exe", "PyMOLWin.exe", "PyMOL.exe", "pymolwin"):
            found = shutil.which(name)
            if found:
                return found
        candidates: list[Path] = []
        roots = [
            os.environ.get("LOCALAPPDATA", ""),
            os.environ.get("ProgramFiles", ""),
            os.environ.get("ProgramFiles(x86)", ""),
            str(Path.home()),
        ]
        for root in roots:
            if not root:
                continue
            base = Path(root)
            candidates += list(base.glob("Schrodinger/PyMOL*/PyMOLWin.exe"))
            candidates += list(base.glob("Schrodinger/PyMOL*/PyMOL.exe"))
            candidates += list(base.glob("PyMOL*/PyMOLWin.exe"))
            candidates += list(base.glob("PyMOL*/PyMOLWin.exe"))
            candidates += list(base.glob("DeLano Scientific/PyMOL*/PyMOLWin.exe"))
        for candidate in candidates:
            if candidate.exists() and candidate.is_file():
                return str(candidate)
        return None

    def open_comparison_dialog(self) -> None:
        """Open pose/scoring comparison tools."""
        if not self.results:
            QMessageBox.information(
                self,
                I18n.get("compare_button", self.lang),
                I18n.get("rt_no_poses_compare", self.lang),
            )
            return
        ComparisonDialog(self.results, self, self.lang).exec()

    def show_validation_report(
        self, rows: list[dict], reference_path: Path, top_n: int
    ) -> None:
        """Open the protocol-validation report dialog."""
        if not rows:
            QMessageBox.warning(
                self,
                I18n.get("validate_protocol", self.lang),
                I18n.get("rt_validation_no_poses", self.lang),
            )
            return
        ProtocolValidationDialog(rows, reference_path, top_n, self, self.lang).exec()

    def update_box_preview(self, box: dict) -> None:
        """Store the current box parameters (viewer removed in v1.1)."""
        self.current_box = dict(box)

    def update_receptor_preview(self, receptor_path: "Path | None") -> None:
        """Store the receptor path for PyMOL visualization."""
        self.current_receptor_path = (
            receptor_path if receptor_path and receptor_path.exists() else None
        )

    def _selected_result_row(self) -> dict | None:
        """Return the stored result row matching the current table selection."""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            return None
        item = self.table.item(selected_row, 0)
        if item is None:
            return None
        row_id = item.data(Qt.UserRole)
        return next((row for row in self.results if self._row_id(row) == row_id), None)

    def load_selected_pose_preview(self) -> None:
        """Load the selected pose into the Py3Dmol preview and interaction panel."""
        row = self._selected_result_row()
        if row is None:
            return
        self._load_pose_preview(row)

    def _load_pose_preview(self, row: dict) -> None:
        """Load one result row into the interactions table and 3D viewer."""
        try:
            view_dir = Path(tempfile.mkdtemp(prefix="vinalab_preview_"))
            receptor_pdb, pose_pdb = prepare_pose_view_files(row, view_dir)
            interactions = self._compute_pose_interactions(row, receptor_pdb, pose_pdb)
            self._populate_interaction_table(interactions)
            if self.preview_view is not None:
                self.preview_view.setHtml(
                    build_pose_view_html(row, receptor_pdb, pose_pdb)
                )
            self.current_preview_row = dict(row)
            self.current_receptor_pdb = receptor_pdb
            self.current_pose_pdb = pose_pdb
            note = self._row_note(row)
            self.pose_detail_label.setText(
                f"Ligante: {row['ligand_name']} | Score: {float(row.get('affinity', 0.0)):.3f} kcal/mol | "
                f"Rank: {int(row.get('pose_rank', row.get('mode', 0)))} | Notas: {note or '-'}"
            )
        except Exception as exc:  # noqa: BLE001 - preview errors should not interrupt result browsing
            self.preview_status.setText(
                f"Erro ao carregar visualização 3D. Verifique os arquivos de entrada. {exc}"
            )

    def export_filtered_subset(self) -> None:
        """Export currently visible result rows to CSV."""
        if not self.filtered_results:
            QMessageBox.information(
                self,
                I18n.get("export_filtered_subset", self.lang),
                I18n.get("rt_no_visible_rows", self.lang),
            )
            return
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            I18n.get("export_filtered_subset", self.lang),
            "",
            I18n.get("rd_csv_filter", self.lang),
        )
        if file_name:
            self._dataframe(self.filtered_results).to_csv(Path(file_name), index=False)

    def export_interaction_table(self) -> None:
        """Export interactions for the selected pose to CSV."""
        if not self.current_interactions:
            QMessageBox.information(
                self,
                I18n.get("export_interaction_table", self.lang),
                I18n.get("rt_no_interactions", self.lang),
            )
            return
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            I18n.get("export_interaction_table", self.lang),
            "",
            I18n.get("rd_csv_filter", self.lang),
        )
        if not file_name:
            return
        rows = [
            {
                "ligand": item["ligand"],
                "pose_rank": item["pose_rank"],
                "residue_name": item["residue_name"],
                "residue_number": item["residue_number"],
                "interaction_type": item["interaction_type"],
                "distance_Å": item["distance_A"],
                "frequency_top10": item["frequency_top10"],
            }
            for item in self.current_interactions
        ]
        pd.DataFrame(rows).to_csv(Path(file_name), index=False)

    def _refresh_interactions_for_selected_pose(self) -> None:
        """Recompute interaction analysis when the cutoff selector changes."""
        if self._selected_result_row() is not None:
            self.load_selected_pose_preview()

    def _populate_interaction_table(self, interactions: list[dict]) -> None:
        """Render the selected pose interaction rows."""
        self.current_interactions = interactions
        self.current_interaction_row = self._selected_result_row()
        self.interaction_table.setRowCount(len(interactions))
        for row_index, item in enumerate(interactions):
            values = [
                f"{item['residue_name']} {item['residue_number']}",
                item["interaction_type"],
                item.get("donor", ""),
                item.get("acceptor", ""),
                f"{float(item['distance_A']):.2f}"
                if item.get("distance_A") != ""
                else "",
                item.get("angle", ""),
                f"{float(item['frequency_top10']):.2f}",
            ]
            for column_index, value in enumerate(values):
                self.interaction_table.setItem(
                    row_index, column_index, QTableWidgetItem(str(value))
                )

    def _compute_pose_interactions(
        self, row: dict, receptor_pdb: Path, pose_pdb: Path
    ) -> list[dict]:
        """Compute receptor-ligand interactions for a selected pose using MDAnalysis."""
        try:
            import MDAnalysis as mda
            import numpy as np
            from MDAnalysis.lib.distances import distance_array
        except ImportError as exc:
            self.preview_status.setText(
                f"MDAnalysis está indisponível; interações ignoradas: {exc}"
            )
            return []

        cutoff = float(self.interaction_cutoff.currentText() or 4.0)
        receptor = mda.Universe(str(receptor_pdb))
        ligand = mda.Universe(str(pose_pdb))
        receptor_atoms = receptor.select_atoms("not name H*")
        ligand_atoms = ligand.select_atoms("not name H*")
        if len(receptor_atoms) == 0 or len(ligand_atoms) == 0:
            return []

        distances = distance_array(receptor_atoms.positions, ligand_atoms.positions)
        interactions_by_key: dict[tuple[str, str, str], dict] = {}
        frequency_by_residue = self._contact_frequency_top10(row, cutoff)

        def add_interaction(
            rec_atom,
            lig_atom,
            interaction_type: str,
            distance: float,
            donor: str = "",
            acceptor: str = "",
            angle: str = "",
        ) -> None:
            residue_key = self._residue_key(rec_atom)
            key = (residue_key[0], residue_key[1], interaction_type)
            previous = interactions_by_key.get(key)
            if previous is not None and float(previous["distance_A"]) <= float(
                distance
            ):
                return
            interactions_by_key[key] = {
                "ligand": row["ligand_name"],
                "pose_rank": int(row.get("pose_rank", row.get("mode", 0))),
                "residue_name": residue_key[0],
                "residue_number": residue_key[1],
                "interaction_type": interaction_type,
                "distance_A": round(float(distance), 3),
                "frequency_top10": round(
                    float(
                        frequency_by_residue.get((residue_key[0], residue_key[1]), 0.0)
                    ),
                    3,
                ),
                "donor": donor,
                "acceptor": acceptor,
                "angle": angle,
            }

        for rec_index, lig_index in np.argwhere(distances <= cutoff):
            rec_atom = receptor_atoms[int(rec_index)]
            lig_atom = ligand_atoms[int(lig_index)]
            add_interaction(
                rec_atom, lig_atom, "Contact", float(distances[rec_index, lig_index])
            )

        for rec_index, lig_index in np.argwhere(distances <= 4.0):
            rec_atom = receptor_atoms[int(rec_index)]
            lig_atom = ligand_atoms[int(lig_index)]
            if (
                self._atom_element(rec_atom) == "C"
                and self._atom_element(lig_atom) == "C"
            ):
                add_interaction(
                    rec_atom,
                    lig_atom,
                    "Hydrophobic",
                    float(distances[rec_index, lig_index]),
                )

        polar_elements = {"N", "O", "S"}
        for rec_index, lig_index in np.argwhere(distances <= 3.5):
            rec_atom = receptor_atoms[int(rec_index)]
            lig_atom = ligand_atoms[int(lig_index)]
            if (
                self._atom_element(rec_atom) not in polar_elements
                or self._atom_element(lig_atom) not in polar_elements
            ):
                continue
            angle = self._estimate_hbond_angle(rec_atom, lig_atom)
            add_interaction(
                rec_atom,
                lig_atom,
                "H-bond",
                float(distances[rec_index, lig_index]),
                donor=self._atom_label(lig_atom),
                acceptor=self._atom_label(rec_atom),
                angle=angle,
            )

        return sorted(
            interactions_by_key.values(),
            key=lambda item: (
                item["residue_number"],
                item["interaction_type"],
                float(item["distance_A"]),
            ),
        )

    def _contact_frequency_top10(
        self, row: dict, cutoff: float
    ) -> dict[tuple[str, str], float]:
        """Return residue contact frequency across the top 10 poses of the same ligand."""
        same_ligand = [
            item
            for item in self.results
            if item.get("ligand_name") == row.get("ligand_name")
        ]
        top_rows = sorted(same_ligand, key=self._docking_score)[:10]
        if not top_rows:
            return {}
        counts: dict[tuple[str, str], int] = {}
        attempted = 0
        for top_row in top_rows:
            try:
                view_dir = Path(tempfile.mkdtemp(prefix="vinalab_contacts_"))
                receptor_pdb, pose_pdb = prepare_pose_view_files(top_row, view_dir)
                for residue_key in self._contact_residue_keys(
                    receptor_pdb, pose_pdb, cutoff
                ):
                    counts[residue_key] = counts.get(residue_key, 0) + 1
                attempted += 1
            except Exception:
                continue
        if attempted == 0:
            return {}
        return {key: value / attempted for key, value in counts.items()}

    def _contact_residue_keys(
        self, receptor_pdb: Path, pose_pdb: Path, cutoff: float
    ) -> set[tuple[str, str]]:
        """Return receptor residues within cutoff of a ligand pose."""
        import MDAnalysis as mda
        import numpy as np
        from MDAnalysis.lib.distances import distance_array

        receptor = mda.Universe(str(receptor_pdb))
        ligand = mda.Universe(str(pose_pdb))
        receptor_atoms = receptor.select_atoms("not name H*")
        ligand_atoms = ligand.select_atoms("not name H*")
        if len(receptor_atoms) == 0 or len(ligand_atoms) == 0:
            return set()
        distances = distance_array(receptor_atoms.positions, ligand_atoms.positions)
        return {
            self._residue_key(receptor_atoms[int(rec_index)])
            for rec_index, _lig_index in np.argwhere(distances <= cutoff)
        }

    def _estimate_hbond_angle(self, donor_atom, acceptor_atom) -> str:
        """Estimate a D-H-A angle when explicit hydrogens are present."""
        try:
            import numpy as np
        except ImportError:
            return "N/A"
        hydrogens = [
            atom for atom in donor_atom.residue.atoms if self._atom_element(atom) == "H"
        ]
        best_angle: float | None = None
        for hydrogen in hydrogens:
            donor_distance = float(
                np.linalg.norm(hydrogen.position - donor_atom.position)
            )
            if donor_distance > 1.25:
                continue
            vector_donor = donor_atom.position - hydrogen.position
            vector_acceptor = acceptor_atom.position - hydrogen.position
            denominator = float(
                np.linalg.norm(vector_donor) * np.linalg.norm(vector_acceptor)
            )
            if denominator == 0:
                continue
            cosine = float(
                np.clip(np.dot(vector_donor, vector_acceptor) / denominator, -1.0, 1.0)
            )
            angle = float(np.degrees(np.arccos(cosine)))
            best_angle = angle if best_angle is None else max(best_angle, angle)
        return f"{best_angle:.1f}" if best_angle is not None else "N/A"

    @staticmethod
    def _atom_element(atom) -> str:
        """Best-effort atom element extraction for PDB/PDBQT-derived files."""
        element = str(getattr(atom, "element", "") or "").strip()
        if element:
            return element.upper()[:1]
        name = str(getattr(atom, "name", "") or "").strip()
        letters = "".join(char for char in name if char.isalpha())
        return letters.upper()[:1] if letters else ""

    @staticmethod
    def _atom_label(atom) -> str:
        """Return a readable atom label."""
        resname = str(getattr(atom, "resname", "") or "").strip()
        resid = str(getattr(atom, "resid", "") or "").strip()
        name = str(getattr(atom, "name", "") or "").strip()
        return f"{resname}{resid}:{name}" if resname or resid else name

    @staticmethod
    def _residue_key(atom) -> tuple[str, str]:
        """Return residue name/number for a receptor atom."""
        return (
            str(getattr(atom, "resname", "") or "").strip(),
            str(getattr(atom, "resid", "") or "").strip(),
        )

    def _set_consensus_tab_visible(self, visible: bool) -> None:
        """Add or remove the consensus tab from the preview/result side panel."""
        current_index = self._side_tab_index(self.consensus_widget)
        if visible and current_index < 0:
            self.consensus_tab_index = self.preview_tabs.addTab(
                self.consensus_widget, I18n.get("consensus", self.lang)
            )
        elif not visible and current_index >= 0:
            self.preview_tabs.removeTab(current_index)
            self.consensus_tab_index = -1
        else:
            self.consensus_tab_index = current_index
        self._set_consensus_plot_tab_visible(visible)

    def _set_consensus_plot_tab_visible(self, visible: bool) -> None:
        """Add or remove the Consensus Plot sub-tab (Issue 14)."""
        current_index = self._side_tab_index(self.consensus_plot_widget)
        if visible and current_index < 0:
            self.consensus_plot_tab_index = self.preview_tabs.addTab(
                self.consensus_plot_widget, "Consensus Plot"
            )
        elif not visible and current_index >= 0:
            self.preview_tabs.removeTab(current_index)
            self.consensus_plot_tab_index = -1
        else:
            self.consensus_plot_tab_index = current_index

    def _consensus_plot_scoring_labels(self) -> list[str]:
        """Return distinct scoring function labels present in results."""
        labels: list[str] = []
        for row in self.results:
            label = str(row.get("scoring_function") or "")
            if label and label not in labels:
                labels.append(label)
        return labels

    def _refresh_consensus_plot_combos(self) -> None:
        """Refresh X/Y/Z dropdown contents to mirror scoring functions in the run."""
        labels = self._consensus_plot_scoring_labels()
        for combo in (
            self.consensus_plot_x_combo,
            self.consensus_plot_y_combo,
            self.consensus_plot_z_combo,
        ):
            previous = combo.currentText()
            combo.blockSignals(True)
            combo.clear()
            combo.addItems(labels)
            if previous and previous in labels:
                combo.setCurrentText(previous)
            combo.blockSignals(False)
        if len(labels) >= 2:
            if not self.consensus_plot_x_combo.currentText():
                self.consensus_plot_x_combo.setCurrentIndex(0)
            if (
                self.consensus_plot_y_combo.currentText()
                == self.consensus_plot_x_combo.currentText()
                and len(labels) > 1
            ):
                self.consensus_plot_y_combo.setCurrentIndex(1)

    def _best_score_per_ligand(self, scoring_label: str) -> dict[str, float]:
        """Return best (most negative) affinity per ligand for a given scoring function."""
        best: dict[str, float] = {}
        for row in self.results:
            if str(row.get("scoring_function") or "") != scoring_label:
                continue
            try:
                score = float(row.get("affinity"))
            except (TypeError, ValueError):
                continue
            ligand = str(row.get("ligand_name") or "")
            if ligand not in best or score < best[ligand]:
                best[ligand] = score
        return best

    def _update_consensus_plot(self) -> None:
        """Render the 2D/3D consensus scatter using matplotlib (plotly fallback when available)."""
        self._refresh_consensus_plot_combos()
        labels = self._consensus_plot_scoring_labels()
        if len(labels) < 2:
            self.consensus_plot_figure.clear()
            self.consensus_plot_canvas.draw_idle()
            return
        x_label = self.consensus_plot_x_combo.currentText() or labels[0]
        y_label = self.consensus_plot_y_combo.currentText() or (
            labels[1] if len(labels) > 1 else labels[0]
        )
        z_label = self.consensus_plot_z_combo.currentText() or (
            labels[2] if len(labels) > 2 else ""
        )
        use_3d = bool(
            self.consensus_plot_3d_checkbox.isChecked()
            and z_label
            and z_label in labels
        )

        scores_x = self._best_score_per_ligand(x_label)
        scores_y = self._best_score_per_ligand(y_label)
        scores_z = self._best_score_per_ligand(z_label) if use_3d else {}
        ligands = sorted(
            set(scores_x)
            & set(scores_y)
            & (set(scores_z) if use_3d else set(scores_x) & set(scores_y))
        )
        if not ligands:
            self.consensus_plot_figure.clear()
            self.consensus_plot_canvas.draw_idle()
            return

        if use_3d:
            try:
                from mpl_toolkits.mplot3d import Axes3D  # noqa: F401 - registers 3D projection
            except ImportError:
                use_3d = False

        self.consensus_plot_figure.clear()
        if use_3d:
            axes = self.consensus_plot_figure.add_subplot(111, projection="3d")
            xs = [scores_x[name] for name in ligands]
            ys = [scores_y[name] for name in ligands]
            zs = [scores_z[name] for name in ligands]
            axes.scatter(xs, ys, zs)
            for name, x, y, z in zip(ligands, xs, ys, zs):
                axes.text(x, y, z, name, fontsize=7)
            axes.set_xlabel(x_label)
            axes.set_ylabel(y_label)
            axes.set_zlabel(z_label)
        else:
            axes = self.consensus_plot_figure.add_subplot(111)
            xs = [scores_x[name] for name in ligands]
            ys = [scores_y[name] for name in ligands]
            scatter = axes.scatter(xs, ys)
            for name, x, y in zip(ligands, xs, ys):
                axes.annotate(
                    name, (x, y), fontsize=7, xytext=(4, 4), textcoords="offset points"
                )
            axes.set_xlabel(x_label)
            axes.set_ylabel(y_label)
            try:
                import mplcursors

                cursor = mplcursors.cursor(scatter, hover=True)
                cursor.connect(
                    "add",
                    lambda sel: sel.annotation.set_text(
                        f"{ligands[sel.index]}\n{x_label}: {xs[sel.index]:.2f}\n{y_label}: {ys[sel.index]:.2f}"
                    ),
                )
            except ImportError:
                pass
        axes.set_title(f"Consensus Plot ({len(ligands)} ligante(s))")
        self.consensus_plot_canvas.draw_idle()

    def _set_clusters_tab_visible(self, visible: bool) -> None:
        """Add or remove the clusters tab from the preview/result side panel."""
        current_index = self._side_tab_index(self.clusters_widget)
        if visible and current_index < 0:
            self.cluster_tab_index = self.preview_tabs.addTab(
                self.clusters_widget, I18n.get("clusters", self.lang)
            )
        elif not visible and current_index >= 0:
            self.preview_tabs.removeTab(current_index)
            self.cluster_tab_index = -1
        else:
            self.cluster_tab_index = current_index

    def _side_tab_index(self, widget: QWidget) -> int:
        """Return the current index of a side-panel tab widget."""
        for index in range(self.preview_tabs.count()):
            if self.preview_tabs.widget(index) is widget:
                return index
        return -1

    def _update_consensus_tab(self) -> None:
        """Compute consensus ranking when two or more scoring functions are present."""
        scoring_labels = self._scoring_headers()
        if len(scoring_labels) < 2:
            self._set_consensus_tab_visible(False)
            return

        self._set_consensus_tab_visible(True)
        consensus_rows = self._build_consensus_rows(scoring_labels)
        headers = [
            I18n.get("ligand_col", self.lang),
            I18n.get("pose_rank", self.lang),
            *scoring_labels,
            I18n.get("mean_rank", self.lang),
            I18n.get("borda_count", self.lang),
            I18n.get("zscore_consensus", self.lang),
            I18n.get("rank_sd", self.lang),
            I18n.get("divergence_flag", self.lang),
        ]
        self.consensus_table.setSortingEnabled(False)
        self.consensus_table.clear()
        self.consensus_table.setColumnCount(len(headers))
        self.consensus_table.setHorizontalHeaderLabels(headers)
        apply_header_tooltips(self.consensus_table)
        self.consensus_table.setRowCount(len(consensus_rows))
        for row_index, item in enumerate(consensus_rows):
            background = self._consensus_color(row_index, len(consensus_rows))
            values = [
                item["ligand"],
                item["pose_rank"],
                *[item["scores"].get(label, "") for label in scoring_labels],
                item["mean_rank"],
                item["borda_count"],
                item["zscore_consensus"],
                item["rank_sd"],
                item["divergence_flag"],
            ]
            for column_index, value in enumerate(values):
                table_item = QTableWidgetItem(str(value))
                if isinstance(value, (int, float)):
                    table_item.setData(Qt.DisplayRole, round(float(value), 3))
                table_item.setBackground(background)
                self.consensus_table.setItem(row_index, column_index, table_item)
        self.consensus_table.setSortingEnabled(True)

    def _build_consensus_rows(self, scoring_labels: list[str]) -> list[dict]:
        """Build consensus metrics across active scoring functions."""
        pose_scores: dict[tuple[str, int], dict] = {}
        for row in self.results:
            scoring_label = row.get("scoring_function", row.get("scoring_key", ""))
            if scoring_label not in scoring_labels:
                continue
            pose_key = self._pose_identity(row)
            entry = pose_scores.setdefault(
                pose_key,
                {"ligand": pose_key[0], "pose_rank": pose_key[1], "scores": {}},
            )
            score = float(row.get("affinity", 0.0))
            previous_score = entry["scores"].get(scoring_label)
            if previous_score is None or score < float(previous_score):
                entry["scores"][scoring_label] = score

        rank_by_label: dict[str, dict[tuple[str, int], int]] = {}
        zscore_by_label: dict[str, dict[tuple[str, int], float]] = {}
        pose_count = len(pose_scores)
        for label in scoring_labels:
            score_rows = [
                (pose_key, data["scores"][label])
                for pose_key, data in pose_scores.items()
                if label in data["scores"]
            ]
            score_rows.sort(key=lambda item: float(item[1]))
            rank_by_label[label] = {
                pose_key: rank
                for rank, (pose_key, _score) in enumerate(score_rows, start=1)
            }
            values = [float(score) for _pose_key, score in score_rows]
            mean_value = statistics.mean(values) if values else 0.0
            stdev_value = statistics.pstdev(values) if len(values) > 1 else 1.0
            if stdev_value == 0:
                stdev_value = 1.0
            zscore_by_label[label] = {
                pose_key: (float(score) - mean_value) / stdev_value
                for pose_key, score in score_rows
            }

        consensus_rows: list[dict] = []
        for pose_key, data in pose_scores.items():
            available_labels = [
                label for label in scoring_labels if pose_key in rank_by_label[label]
            ]
            if len(available_labels) < 2:
                continue
            ranks = [rank_by_label[label][pose_key] for label in available_labels]
            mean_rank = statistics.mean(ranks)
            borda_count = sum(pose_count - rank for rank in ranks)
            zscore_consensus = sum(
                zscore_by_label[label][pose_key] for label in available_labels
            )
            rank_sd = statistics.pstdev(ranks) if len(ranks) > 1 else 0.0
            consensus_rows.append(
                {
                    "ligand": data["ligand"],
                    "pose_rank": data["pose_rank"],
                    "scores": data["scores"],
                    "mean_rank": round(float(mean_rank), 3),
                    "borda_count": int(borda_count),
                    "zscore_consensus": round(float(zscore_consensus), 3),
                    "rank_sd": round(float(rank_sd), 3),
                    "divergence_flag": "Robust"
                    if rank_sd <= 1.0
                    else "⚠ Controversial",
                }
            )
        return sorted(
            consensus_rows,
            key=lambda item: (
                float(item["mean_rank"]),
                float(item["zscore_consensus"]),
            ),
        )

    @staticmethod
    def _pose_identity(row: dict) -> tuple[str, int]:
        """Return the grouping key used for consensus metrics."""
        return (
            str(row.get("ligand_name", "")),
            int(row.get("pose_rank", row.get("mode", 0))),
        )

    @staticmethod
    def _consensus_color(row_index: int, row_count: int) -> QColor:
        """Return quartile color for consensus ranking rows."""
        if row_count <= 1:
            return QColor("#d8ead3")
        percentile = row_index / max(row_count - 1, 1)
        if percentile <= 0.25:
            return QColor("#d8ead3")
        if percentile >= 0.75:
            return QColor("#f2c6bd")
        return QColor("#f5e6a6")

    def _cluster_cutoff_changed(self, value: int) -> None:
        """Update RMSD cluster cutoff from the slider."""
        self.cluster_cutoff_label.setText(f"RMSD cutoff: {value / 100:.1f} Å")
        self._update_clusters_tab()

    def _update_clusters_tab(self) -> None:
        """Compute and render RMSD pose clusters."""
        if not self.results:
            self._set_clusters_tab_visible(False)
            return
        self._set_clusters_tab_visible(True)
        try:
            self.current_clusters = self._build_pose_clusters(
                self.cluster_cutoff_slider.value() / 100
            )
        except ImportError as exc:
            self.current_clusters = []
            self.cluster_table.setRowCount(1)
            self.cluster_table.setColumnCount(1)
            self.cluster_table.setHorizontalHeaderLabels(["Clusters"])
            apply_header_tooltips(self.cluster_table)
            self.cluster_table.setItem(
                0, 0, QTableWidgetItem(f"scipy está indisponível: {exc}")
            )
            return
        self._populate_cluster_table()
        self._update_cluster_histogram()

    def _build_pose_clusters(self, cutoff: float) -> list[dict]:
        """Cluster poses of each ligand by heavy-atom RMSD (delegates to results_clustering)."""
        return build_pose_clusters(
            self.results,
            self._pose_heavy_coordinates,
            self._pose_rmsd,
            self._docking_score,
            cutoff,
        )

    def _populate_cluster_table(self) -> None:
        """Render RMSD clusters into the cluster table."""
        headers = [
            I18n.get("ligand_col", self.lang),
            I18n.get("cluster_id", self.lang),
            I18n.get("cluster_size", self.lang),
            I18n.get("best_score", self.lang),
            I18n.get("representative_pose", self.lang),
            I18n.get("members", self.lang),
        ]
        self.cluster_table.clear()
        self.cluster_table.setColumnCount(len(headers))
        self.cluster_table.setHorizontalHeaderLabels(headers)
        apply_header_tooltips(self.cluster_table)
        self.cluster_table.setRowCount(len(self.current_clusters))
        for row_index, cluster in enumerate(self.current_clusters):
            representative = cluster["representative"]
            values = [
                cluster["ligand"],
                cluster["cluster_id"],
                cluster["size"],
                round(float(cluster["best_score"]), 3),
                int(representative.get("pose_rank", representative.get("mode", 0))),
                ", ".join(
                    str(int(member.get("pose_rank", member.get("mode", 0))))
                    for member in cluster["members"]
                ),
            ]
            for column_index, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if isinstance(value, (int, float)):
                    item.setData(Qt.DisplayRole, value)
                item.setData(Qt.UserRole, self._row_id(representative))
                self.cluster_table.setItem(row_index, column_index, item)

    def _update_cluster_histogram(self) -> None:
        """Draw the cluster size histogram."""
        self.cluster_figure.clear()
        axis = self.cluster_figure.add_subplot(111)
        if self.current_clusters:
            labels = [
                f"{cluster['ligand']} C{cluster['cluster_id']}"
                for cluster in self.current_clusters
            ]
            sizes = [int(cluster["size"]) for cluster in self.current_clusters]
            axis.bar(range(len(sizes)), sizes, color="#4f7cac")
            axis.set_xticks(range(len(labels)))
            axis.set_xticklabels(labels, rotation=45, ha="right", fontsize=8)
            axis.set_ylabel("Tamanho do cluster")
        self.cluster_canvas.draw_idle()
        if self.cluster_web is not None and self.current_clusters:
            try:
                html_path = cluster_chart_html(
                    self.current_clusters, I18n.get("cluster_size", self.lang)
                )
                self.cluster_web.setUrl(QUrl.fromLocalFile(str(html_path)))
            except Exception as exc:  # noqa: BLE001 - chart is non-critical
                logger.warning(
                    "Falha ao renderizar gráfico Plotly de clusters: %s", exc
                )

    def _cluster_row_clicked(self, row_index: int, _column_index: int) -> None:
        """Load a cluster representative when a cluster row is clicked."""
        if row_index < 0 or row_index >= len(self.current_clusters):
            return
        representative = self.current_clusters[row_index]["representative"]
        row_id = self._row_id(representative)
        for table_row in range(self.table.rowCount()):
            item = self.table.item(table_row, 0)
            if item is not None and item.data(Qt.UserRole) == row_id:
                self.table.selectRow(table_row)
                break
        self._load_pose_preview(representative)

    def export_cluster_representatives(self) -> None:
        """Export the representative pose of each RMSD cluster."""
        if not self.current_clusters:
            QMessageBox.information(
                self,
                I18n.get("export_best_cluster", self.lang),
                I18n.get("rt_no_clusters_export", self.lang),
            )
            return
        export_format, ok = QInputDialog.getItem(
            self,
            I18n.get("export_best_cluster", self.lang),
            I18n.get("rd_format_label", self.lang),
            ["pdbqt", "pdb", "mol2"],
            0,
            False,
        )
        if not ok:
            return
        output_folder = QFileDialog.getExistingDirectory(
            self, I18n.get("select_output_dir", self.lang)
        )
        if not output_folder:
            return
        if export_format in {"pdb", "mol2"} and find_obabel_executable() is None:
            QMessageBox.critical(
                self,
                I18n.get("export_best_cluster", self.lang),
                I18n.get("rd_obabel_missing_export", self.lang),
            )
            return
        output_dir = Path(output_folder)
        exported = 0
        for cluster in self.current_clusters:
            representative = cluster["representative"]
            basename = f"{safe_export_name(cluster['ligand'])}_cluster{int(cluster['cluster_id'])}_pose{int(representative['mode'])}"
            output_path = output_dir / f"{basename}.{export_format}"
            self._export_pose_to_path(representative, output_path, export_format)
            exported += 1
        QMessageBox.information(
            self,
            I18n.get("export_best_cluster", self.lang),
            I18n.get("rt_cluster_exported", self.lang).format(
                count=exported, directory=output_dir
            ),
        )

    def _export_pose_to_path(
        self, row: dict, output_path: Path, export_format: str
    ) -> None:
        """Write one pose row to a target path, converting with Open Babel if needed."""
        temporary_pdbqt = (
            output_path
            if export_format == "pdbqt"
            else output_path.with_suffix(".pdbqt")
        )
        temporary_pdbqt.write_text(
            extract_pose_model(
                Path(row["output_file"]), int(row["mode"]), include_model=False
            ),
            encoding="utf-8",
        )
        if export_format == "pdbqt":
            return
        convert_with_obabel(temporary_pdbqt, output_path)
        temporary_pdbqt.unlink(missing_ok=True)

    def _pose_text(self, row: dict) -> str:
        """Return the single-model PDBQT text for one result row's pose."""
        return extract_pose_model(
            Path(row["output_file"]), int(row["mode"]), include_model=False
        )

    def _pose_heavy_coordinates(self, row: dict):
        """Extract heavy-atom coordinates (file order) from one PDBQT pose model."""
        return pdbqt_heavy_coordinates(self._pose_text(row))

    @staticmethod
    def _pose_rmsd(left_coordinates, right_coordinates) -> float:
        """Direct heavy-atom RMSD between two poses that share an atom ordering.

        Valid only for poses of the same ligand from one Vina run (identical atom
        order). Raises ValueError on shape mismatch instead of returning a
        misleading sentinel; cross-ligand or cross-reference comparisons must use
        ``core.rmsd.symmetry_corrected_rmsd`` instead.
        """
        return pose_pair_rmsd(left_coordinates, right_coordinates)

    def _apply_filters(self) -> None:
        """Apply all visible filters and re-render the table."""
        rows = [row for row in self.results if self._row_matches_filters(row)]
        if self._sort_criteria:
            rows = self._sort_rows(rows)
        self.filtered_results = rows
        self._populate_table()
        self._update_chart()
        self._save_state()

    def _row_matches_filters(self, row: dict) -> bool:
        """Return True when a result row passes all active filters."""
        ligand_query = self.ligand_filter.text().strip().lower()
        if ligand_query and ligand_query not in row["ligand_name"].lower():
            return False
        selected_scoring = self._selected_scoring_filters()
        scoring_label = row.get("scoring_function", row.get("scoring_key", ""))
        if selected_scoring and scoring_label not in selected_scoring:
            return False
        docking_score = self._docking_score(row)
        if (
            docking_score < self.affinity_min.value()
            or docking_score > self.affinity_max.value()
        ):
            return False
        rmsd = float(row.get("rmsd_lb", 0.0))
        if rmsd < self.rmsd_min.value() or rmsd > self.rmsd_max.value():
            return False
        if self.pinned_only.isChecked() and not self._row_pinned(row):
            return False
        return True

    def _table_values(self, row: dict, scoring_headers: list[str]) -> list:
        """Return display values for one row in the enhanced results table."""
        scoring_label = row.get("scoring_function", row.get("scoring_key", ""))
        scoring_values = [
            row.get("affinity", "") if header == scoring_label else ""
            for header in scoring_headers
        ]
        return [
            row["ligand_name"],
            row.get("pose_rank", row.get("mode", "")),
            self._docking_score(row),
            *scoring_values,
            row.get("rmsd_lb", ""),
            self._row_pinned(row),
            self._row_note(row),
            row.get("scoring_error", ""),
        ]

    def _table_item_changed(self, item: QTableWidgetItem) -> None:
        """Persist pin and note edits made directly in the table."""
        if self._rendering_table or item is None:
            return
        row_id = item.data(Qt.UserRole)
        if not row_id:
            return
        key = (
            self.column_keys[item.column()]
            if item.column() < len(self.column_keys)
            else ""
        )
        if key == "pinned":
            self.table_state.setdefault("pins", {})[row_id] = (
                item.checkState() == Qt.Checked
            )
        elif key == "notes":
            self.table_state.setdefault("notes", {})[row_id] = item.text()
        else:
            return
        self._save_state()

    def _open_context_menu(self, position: QPoint) -> None:
        """Open the result-row context menu."""
        item = self.table.itemAt(position)
        if item is None:
            return
        self.table.selectRow(item.row())
        row = self._selected_result_row()
        if row is None:
            return
        menu = QMenu(self)
        pin_action = QAction("Unpin" if self._row_pinned(row) else "Pin", self)
        note_action = QAction("Adicionar nota", self)
        export_action = QAction("Exportar esta pose", self)
        pin_action.triggered.connect(lambda: self._toggle_selected_pin(row))
        note_action.triggered.connect(lambda: self._edit_note(row))
        export_action.triggered.connect(lambda: self._export_single_pose(row))
        menu.addAction(pin_action)
        menu.addAction(note_action)
        menu.addAction(export_action)
        menu.exec(self.table.viewport().mapToGlobal(position))

    def _toggle_selected_pin(self, row: dict) -> None:
        """Toggle pin status for a result row."""
        row_id = self._row_id(row)
        self.table_state.setdefault("pins", {})[row_id] = not self._row_pinned(row)
        self._save_state()
        self._apply_filters()

    def _edit_note(self, row: dict) -> None:
        """Prompt for and persist a note for a result row."""
        row_id = self._row_id(row)
        note, ok = QInputDialog.getText(
            self, "Nota da pose", "Notas", text=self._row_note(row)
        )
        if ok:
            self.table_state.setdefault("notes", {})[row_id] = note
            self._save_state()
            self._apply_filters()

    def _export_single_pose(self, row: dict) -> None:
        """Export one selected pose from the context menu."""
        default_name = (
            f"{safe_export_name(row['ligand_name'])}_pose{int(row['mode'])}.pdbqt"
        )
        file_name, selected_filter = QFileDialog.getSaveFileName(
            self,
            I18n.get("rt_export_this_pose", self.lang),
            default_name,
            I18n.get("rt_pose_export_filter", self.lang),
        )
        if not file_name:
            return
        output_path = Path(file_name)
        pose_pdbqt = (
            output_path
            if output_path.suffix.lower() == ".pdbqt"
            else output_path.with_suffix(".pdbqt")
        )
        pose_pdbqt.write_text(
            extract_pose_model(
                Path(row["output_file"]), int(row["mode"]), include_model=False
            ),
            encoding="utf-8",
        )
        if output_path.suffix.lower() != ".pdbqt":
            if find_obabel_executable() is None:
                QMessageBox.critical(
                    self,
                    I18n.get("rt_export_this_pose", self.lang),
                    I18n.get("rd_obabel_missing_export", self.lang),
                )
                return
            convert_with_obabel(pose_pdbqt, output_path)
            pose_pdbqt.unlink(missing_ok=True)
        QMessageBox.information(
            self,
            I18n.get("rt_export_this_pose", self.lang),
            I18n.get("rt_exported_to", self.lang).format(path=output_path),
        )

    def _header_clicked(self, section: int) -> None:
        """Sort rows by one column, preserving earlier keys with Shift-click."""
        if section >= len(self.column_keys):
            return
        key = self.column_keys[section]
        shift = QApplication.keyboardModifiers() & Qt.ShiftModifier
        existing = next((item for item in self._sort_criteria if item[0] == key), None)
        ascending = not existing[1] if existing else True
        if shift:
            self._sort_criteria = [
                item for item in self._sort_criteria if item[0] != key
            ]
            self._sort_criteria.append((key, ascending))
        else:
            self._sort_criteria = [(key, ascending)]
        self._apply_filters()

    def _sort_rows(self, rows: list[dict]) -> list[dict]:
        """Sort rows using the active single or multi-column criteria."""
        sorted_rows = list(rows)
        for key, ascending in reversed(self._sort_criteria):
            sorted_rows.sort(
                key=lambda row: self._sort_value(row, key), reverse=not ascending
            )
        return sorted_rows

    def _sort_value(self, row: dict, key: str):
        """Return a comparable sort value for a row and column key."""
        if key.startswith("score::"):
            label = key.split("::", 1)[1]
            return (
                float(row.get("affinity", 999999.0))
                if row.get("scoring_function") == label
                else 999999.0
            )
        if key == "docking_score":
            return self._docking_score(row)
        if key == "pose_rank":
            return int(row.get("pose_rank", row.get("mode", 0)))
        if key == "rmsd_lb":
            return float(row.get("rmsd_lb", 999999.0))
        if key == "pinned":
            return int(self._row_pinned(row))
        if key == "notes":
            return self._row_note(row).lower()
        return str(row.get(key, "")).lower()

    def _sync_state_path(self) -> None:
        """Load the JSON sidecar for the current output directory, if any."""
        output_dirs = {
            Path(row["output_file"]).parent
            for row in self.results
            if row.get("output_file")
        }
        if not output_dirs:
            return
        output_dir = sorted(output_dirs, key=lambda path: str(path))[0]
        if self._state_loaded_for == output_dir:
            return
        self.state_path = output_dir / "vina_results_state.json"
        self._state_loaded_for = output_dir
        if self.state_path.exists():
            try:
                self.table_state = json.loads(
                    self.state_path.read_text(encoding="utf-8")
                )
            except json.JSONDecodeError:
                self.table_state = {"pins": {}, "notes": {}, "filters": {}}
        self._restore_filters()

    def _save_state(self) -> None:
        """Persist pins, notes, and filter state beside the docking output."""
        if self.state_path is None:
            return
        self.table_state["filters"] = self._current_filter_state()
        self.state_path.write_text(
            json.dumps(self.table_state, indent=2, ensure_ascii=False), encoding="utf-8"
        )

    def _restore_filters(self) -> None:
        """Restore saved filter controls from sidecar state."""
        filters = self.table_state.get("filters", {})
        self.ligand_filter.blockSignals(True)
        self.affinity_min.blockSignals(True)
        self.affinity_max.blockSignals(True)
        self.rmsd_min.blockSignals(True)
        self.rmsd_max.blockSignals(True)
        self.pinned_only.blockSignals(True)
        self.ligand_filter.setText(filters.get("ligand", ""))
        self.affinity_min.setValue(float(filters.get("affinity_min", -9999.0)))
        self.affinity_max.setValue(float(filters.get("affinity_max", 9999.0)))
        self.rmsd_min.setValue(float(filters.get("rmsd_min", 0.0)))
        self.rmsd_max.setValue(float(filters.get("rmsd_max", 9999.0)))
        self.pinned_only.setChecked(bool(filters.get("pinned_only", False)))
        self.ligand_filter.blockSignals(False)
        self.affinity_min.blockSignals(False)
        self.affinity_max.blockSignals(False)
        self.rmsd_min.blockSignals(False)
        self.rmsd_max.blockSignals(False)
        self.pinned_only.blockSignals(False)

    def _current_filter_state(self) -> dict:
        """Return current filter values for persistence."""
        return {
            "ligand": self.ligand_filter.text(),
            "scoring": list(self._selected_scoring_filters()),
            "affinity_min": self.affinity_min.value(),
            "affinity_max": self.affinity_max.value(),
            "rmsd_min": self.rmsd_min.value(),
            "rmsd_max": self.rmsd_max.value(),
            "pinned_only": self.pinned_only.isChecked(),
        }

    def _refresh_scoring_filter_options(self) -> None:
        """Rebuild the multi-select scoring filter menu from result rows."""
        selected = set(self.table_state.get("filters", {}).get("scoring", []))
        labels = sorted(
            {
                row.get("scoring_function", row.get("scoring_key", ""))
                for row in self.results
            }
        )
        self.scoring_filter_menu.clear()
        self.scoring_filter_actions = {}
        for label in labels:
            action = QAction(label, self)
            action.setCheckable(True)
            action.setChecked(not selected or label in selected)
            action.toggled.connect(lambda _checked: self._scoring_filter_changed())
            self.scoring_filter_menu.addAction(action)
            self.scoring_filter_actions[label] = action
        self._refresh_scoring_filter_button()

    def _scoring_filter_changed(self) -> None:
        """Persist and apply scoring filter changes."""
        self._refresh_scoring_filter_button()
        self._apply_filters()

    def _selected_scoring_filters(self) -> set[str]:
        """Return checked scoring labels; empty means all."""
        if not self.scoring_filter_actions:
            return set()
        checked = {
            label
            for label, action in self.scoring_filter_actions.items()
            if action.isChecked()
        }
        return set() if len(checked) == len(self.scoring_filter_actions) else checked

    def _refresh_scoring_filter_button(self) -> None:
        """Update the scoring filter button label."""
        selected = self._selected_scoring_filters()
        if not selected:
            self.scoring_filter_button.setText(I18n.get("scoring_all", self.lang))
        elif len(selected) == 1:
            self.scoring_filter_button.setText(f"Scoring: {next(iter(selected))}")
        else:
            self.scoring_filter_button.setText(f"Scoring: {len(selected)} selected")

    def _row_id(self, row: dict) -> str:
        """Return a stable sidecar key for a result row."""
        return "|".join(
            [
                str(row.get("ligand_name", "")),
                str(row.get("scoring_key", row.get("scoring_function", ""))),
                str(row.get("mode", "")),
                str(row.get("output_file", "")),
            ]
        )

    def _row_pinned(self, row: dict) -> bool:
        """Return persisted pin status for a row."""
        return bool(self.table_state.get("pins", {}).get(self._row_id(row), False))

    def _row_note(self, row: dict) -> str:
        """Return persisted notes for a row."""
        return str(self.table_state.get("notes", {}).get(self._row_id(row), ""))

    @staticmethod
    def _docking_score(row: dict) -> float:
        """Return the original docking score for a row."""
        return float(row.get("vina_affinity", row.get("affinity", 0.0)) or 0.0)

    @staticmethod
    def _numeric_column(key: str) -> bool:
        """Return True when a table column should sort as numeric."""
        return key in {"pose_rank", "docking_score", "rmsd_lb"} or key.startswith(
            "score::"
        )

    @staticmethod
    def _range_spin(value: float) -> QDoubleSpinBox:
        """Create a compact numeric filter spin box."""
        spin = QDoubleSpinBox()
        spin.setRange(-9999.0, 9999.0)
        spin.setDecimals(2)
        spin.setSingleStep(0.5)
        spin.setValue(value)
        spin.setMinimumWidth(72)
        spin.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
        return spin

    def _scoring_headers(self) -> list[str]:
        """Return dynamic scorer column names."""
        return sorted(
            {
                row.get("scoring_function", row.get("scoring_key", ""))
                for row in self.results
                if row.get("scoring_function") or row.get("scoring_key")
            }
        )

    @staticmethod
    def _color_for_affinity(affinity: float) -> QColor:
        """Return a row color based on affinity score thresholds."""
        if affinity <= -9:
            return QColor("#2f6f4e")
        if affinity <= -7:
            return QColor("#8a7a2e")
        return QColor("#f5f5f5")

    @staticmethod
    def _excel_fill_for_affinity(affinity: float) -> PatternFill:
        """Return an Excel fill color based on affinity score thresholds."""
        if affinity <= -9:
            return PatternFill("solid", fgColor="5BA36A")
        if affinity <= -7:
            return PatternFill("solid", fgColor="E0C95A")
        return PatternFill("solid", fgColor="FFFFFF")
